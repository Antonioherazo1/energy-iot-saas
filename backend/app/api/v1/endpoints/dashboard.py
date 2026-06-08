import csv
import io
import uuid
from datetime import datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import and_, select, text
from sqlalchemy.orm import Session

from app.api.deps import get_current_user
from app.core.config import settings
from app.db.session import get_db
from app.models.device import Device
from app.models.telemetry import Telemetry
from app.models.user import User
from app.schemas.dashboard import DashboardSummaryRead, DeviceStatusRead, EnergyBucketRead, LatestTelemetryRead
from app.services.dashboard_service import get_accessible_organization_ids, get_billing_current_daily, get_billing_daily_per_channel, get_billing_monthly_energy, get_channel_day_series, get_device_status, get_energy_by_period, get_latest_telemetry, get_realtime_currents, get_summary, recalculate_daily_energy

router = APIRouter()


@router.get("/summary", response_model=DashboardSummaryRead)
def dashboard_summary(
    organization_id: uuid.UUID | None = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    return get_summary(db=db, user=current_user, organization_id=organization_id)


@router.get("/devices/status", response_model=list[DeviceStatusRead])
def devices_status(
    organization_id: uuid.UUID | None = None,
    online_window_minutes: int = Query(default=5, ge=1, le=120),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[dict]:
    return get_device_status(
        db=db,
        user=current_user,
        organization_id=organization_id,
        online_window_minutes=online_window_minutes,
    )


@router.get("/telemetry/latest", response_model=list[LatestTelemetryRead])
def telemetry_latest(
    organization_id: uuid.UUID | None = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[dict]:
    return get_latest_telemetry(db=db, user=current_user, organization_id=organization_id)


@router.get("/energy/daily", response_model=list[EnergyBucketRead])
def energy_daily(
    organization_id: uuid.UUID | None = None,
    limit: int = Query(default=30, ge=1, le=365),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[dict]:
    return get_energy_by_period(db=db, user=current_user, organization_id=organization_id, period="day", limit=limit)


@router.get("/channels/latest")
def channels_latest(
    organization_id: uuid.UUID | None = None,
    limit: int = Query(default=60, ge=10, le=500),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[dict]:
    from app.services.dashboard_service import get_channel_time_series
    return get_channel_time_series(db=db, user=current_user, organization_id=organization_id, limit=limit)


@router.get("/channels/day")
def channels_day(
    device_id: uuid.UUID,
    date: str = Query(description="ISO date YYYY-MM-DD"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[dict]:
    return get_channel_day_series(db=db, user=current_user, device_id=device_id, date=date)


@router.get("/channels/realtime")
def channels_realtime(
    device_id: uuid.UUID,
    minutes: int = Query(default=10, ge=1, le=60),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[dict]:
    return get_realtime_currents(db=db, user=current_user, device_id=device_id, minutes=minutes)


@router.get("/energy/monthly", response_model=list[EnergyBucketRead])
def energy_monthly(
    organization_id: uuid.UUID | None = None,
    limit: int = Query(default=12, ge=1, le=120),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[dict]:
    return get_energy_by_period(db=db, user=current_user, organization_id=organization_id, period="month", limit=limit)


@router.get("/energy/billing/monthly")
def energy_billing_monthly(
    billing_start_day: int = Query(default=1, ge=1, le=28),
    limit: int = Query(default=6, ge=1, le=24),
    organization_id: uuid.UUID | None = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[dict]:
    return get_billing_monthly_energy(
        db=db, user=current_user, billing_start_day=billing_start_day, limit=limit, organization_id=organization_id,
    )


@router.get("/energy/billing/daily")
def energy_billing_daily(
    billing_start_day: int = Query(default=1, ge=1, le=28),
    organization_id: uuid.UUID | None = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[dict]:
    return get_billing_current_daily(
        db=db, user=current_user, billing_start_day=billing_start_day, organization_id=organization_id,
    )


@router.get("/energy/billing/daily/channels")
def energy_billing_daily_channels(
    device_id: uuid.UUID,
    organization_id: uuid.UUID | None = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[dict]:
    return get_billing_daily_per_channel(
        db=db, user=current_user, device_id=device_id, organization_id=organization_id,
    )


@router.get("/telemetry/csv")
def telemetry_csv(
    organization_id: uuid.UUID | None = None,
    start: str | None = Query(default=None, description="ISO datetime (inicio)"),
    end: str | None = Query(default=None, description="ISO datetime (fin)"),
    limit: int = Query(default=500, ge=1, le=10000),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    organization_ids = get_accessible_organization_ids(db, current_user, organization_id)
    if not organization_ids:
        return _empty_excel()

    filters = [Device.organization_id.in_(organization_ids)]
    if start:
        filters.append(Telemetry.recorded_at >= datetime.fromisoformat(start))
    if end:
        filters.append(Telemetry.recorded_at <= datetime.fromisoformat(end))

    query = (
        select(Telemetry, Device)
        .join(Device, Telemetry.device_id == Device.id)
        .where(and_(*filters))
        .order_by(Telemetry.recorded_at.desc())
        .limit(limit)
    )
    rows = db.execute(query).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Telemetria"
    bold = Font(bold=True)
    headers = ["Dispositivo", "Codigo", "Fecha", "Voltaje (V)", "Corriente (A)", "Potencia (W)", "Energia (kWh)", "Frecuencia (Hz)", "FP", "CH1 (A)", "CH2 (A)", "CH3 (A)", "CH4 (A)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = bold
    for telemetry, device in rows:
        ws.append([
            device.name,
            device.code,
            telemetry.recorded_at.isoformat() if telemetry.recorded_at else "",
            telemetry.voltage,
            telemetry.current,
            telemetry.power,
            telemetry.energy_kwh,
            telemetry.frequency,
            telemetry.power_factor,
            telemetry.ch1,
            telemetry.ch2,
            telemetry.ch3,
            telemetry.ch4,
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=telemetria.xlsx"},
    )


def _empty_excel() -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Telemetria"
    bold = Font(bold=True)
    headers = ["Dispositivo", "Codigo", "Fecha", "Voltaje (V)", "Corriente (A)", "Potencia (W)", "Energia (kWh)", "Frecuencia (Hz)", "FP", "CH1 (A)", "CH2 (A)", "CH3 (A)", "CH4 (A)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = bold
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=telemetria.xlsx"},
    )


@router.get("/telemetry/range", response_model=list[LatestTelemetryRead])
def telemetry_range(
    organization_id: uuid.UUID | None = None,
    start: str = Query(description="ISO datetime (inicio)"),
    end: str = Query(description="ISO datetime (fin)"),
    limit: int = Query(default=200, ge=1, le=5000),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[dict]:
    organization_ids = get_accessible_organization_ids(db, current_user, organization_id)
    if not organization_ids:
        return []

    filters = [
        Device.organization_id.in_(organization_ids),
        Telemetry.recorded_at >= datetime.fromisoformat(start),
        Telemetry.recorded_at <= datetime.fromisoformat(end),
    ]
    rows = db.execute(
        select(
            Telemetry.device_id,
            Telemetry.recorded_at,
            Telemetry.voltage,
            Telemetry.current,
            Telemetry.power,
            Telemetry.energy_kwh,
            Telemetry.frequency,
            Telemetry.power_factor,
            Telemetry.ch1,
            Telemetry.ch2,
            Telemetry.ch3,
            Telemetry.ch4,
            Device.name.label("device_name"),
            Device.code.label("device_code"),
        )
        .join(Device, Telemetry.device_id == Device.id)
        .where(and_(*filters))
        .order_by(Telemetry.recorded_at.desc())
        .limit(limit)
    )
    return [dict(row._mapping) for row in rows]


@router.get("/db-size")
def db_size(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    result = db.execute(text("SELECT pg_database_size(current_database()) / 1048576.0 AS size_mb"))
    row = result.one()
    return {"size_mb": round(float(row.size_mb), 1)}


@router.post("/energy/recalculate")
def energy_recalculate(
    days: int = Query(default=30, ge=1, le=365),
    organization_id: uuid.UUID | None = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[dict]:
    return recalculate_daily_energy(
        db=db, user=current_user, days=days, organization_id=organization_id,
    )

